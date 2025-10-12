import json
import logging
import threading
import traceback
import datetime as dt

from openpyxl import Workbook
from django.utils import timezone
from django.shortcuts import render
from django.http import HttpResponse
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from dashboard.models import LienData, RealEstateData
from openpyxl.styles import Font, PatternFill, Alignment

# New import from the neutral state file
from dashboard.utils.state import stop_scraper_flag
from dashboard.utils.init_scraper import (run_lien_scraper, 
                                          run_realestate_scraper)


# ------------------ LOGGER SETUP -------------------
logger = logging.getLogger(__name__)


# ------------------DASHBOARD VIEWS -------------------
def dashboard(request):
    lien_data = LienData.objects.all().order_by('-created_at')
    realestate_data = RealEstateData.objects.all().order_by('-created_at')
    return render(request, 'dashboard.html', {
        'lien_data': lien_data,
        'realestate_data': realestate_data
    })
    
    
def start_scraper(request):
    try:
        if request.method == 'POST':
            data = request.POST.dict()
            scraper_type = data.get('scraper_type')
            
            if scraper_type == 'lien':
                # Convert dates
                raw_to_date = request.POST.get("to_date")
                raw_from_date = request.POST.get("from_date")
                to_date_mmddyyyy = dt.date.fromisoformat(raw_to_date).strftime("%m/%d/%Y") if raw_to_date else ""
                from_date_mmddyyyy = dt.date.fromisoformat(raw_from_date).strftime("%m/%d/%Y") if raw_from_date else ""
                data['to_date'] = to_date_mmddyyyy
                data['from_date'] = from_date_mmddyyyy

                thread = threading.Thread(
                    target=run_lien_scraper,
                    kwargs={"params": data},
                    daemon=True,
                )
                msg = 'Lien scraper started'
            elif scraper_type == 'realestate':
                # Convert dates for scraper
                raw_to_date = request.POST.get("txtFromDate")
                raw_from_date = request.POST.get("txtToDate")
                to_date_mmddyyyy = dt.date.fromisoformat(raw_to_date).strftime("%m/%d/%Y") if raw_to_date else ""
                from_date_mmddyyyy = dt.date.fromisoformat(raw_from_date).strftime("%m/%d/%Y") if raw_from_date else ""
                data['txtFromDate'] = to_date_mmddyyyy
                data['txtToDate'] = from_date_mmddyyyy
                # Real estate scraper now accepts parameters from the form
                thread = threading.Thread(
                    target=run_realestate_scraper, 
                    kwargs={"params": data},
                    daemon=True,
                )
                msg = 'Real estate scraper started'
            thread.start()
            return JsonResponse({'status': msg}, status=200)
    except Exception as e:
        logger.error(f"Error starting scraper: {e}\n{traceback.format_exc()}") 
        return JsonResponse({'error': f'Invalid request: \n{e}'}, status=400)


@csrf_exempt
def stop_scraper(request):
    """View to set the global stop flag for a specific scraper type."""
    try:
        if request.method == 'POST':
            # Note: We use json.loads(request.body) for the stopScraper JavaScript fetch call
            data = json.loads(request.body)
            scraper_type = data.get('scraper_type')
            
            if scraper_type == 'lien':
                stop_scraper_flag['lien'] = True
                msg = 'Lien scraper stop signal sent. It will stop after the current step.'
            elif scraper_type == 'realestate':
                stop_scraper_flag['realestate'] = True
                msg = 'Real estate scraper stop signal sent. It will stop after the current step.'
            else:
                return JsonResponse({'error': 'Invalid scraper type'}, status=400)
            
            return JsonResponse({'status': msg}, status=200)
    except Exception as e:
        logger.error(f"Error stopping scraper: {e}\n{traceback.format_exc()}")
        return JsonResponse({'error': f'Invalid request: \n{e}'}, status=400)


def get_latest_data(request):
    data_type = request.GET.get('type', 'lien')
    if data_type == 'lien':
        data = list(LienData.objects.all().order_by('-created_at').values())
    else:
        data = list(RealEstateData.objects.all().order_by('-created_at').values())
    return JsonResponse({'data': data}, status=200)

    
# ------------------ EXCEL DOWNLOAD VIEWS -------------------
@csrf_exempt
def download_lien_excel(request):
    """Download single lien record as Excel"""
    try:
        if request.method == 'POST':
            data = json.loads(request.body)
            pdf_url = data.get('pdf_url', '')
            
            # Find the lien record based on PDF URL
            lien_record = LienData.objects.filter(pdf_document_url=pdf_url).first()
            
            if not lien_record:
                return JsonResponse({'error': 'Record not found'}, status=404)
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Lien Record"
            
            # Add headers with styling
            headers = [
                'Direct Party (Debtor)', 'Reverse Party (Claimant)', 'Address', 
                'Zipcode', 'Total Due', 'County', 'Instrument', 'Date Filed',
                'Book', 'Page', 'Description', 'PDF Document URL', 'PDF File'
            ]
            
            # Style headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Add data
            data_row = [
                lien_record.direct_party_debtor or '',
                lien_record.reverse_party_claimant or '',
                lien_record.address or '',
                lien_record.zipcode or '',
                lien_record.total_due or '',
                lien_record.county or '',
                lien_record.instrument or '',
                lien_record.date_filed or '',
                lien_record.book or '',
                lien_record.page or '',
                lien_record.description or '',
                lien_record.pdf_document_url or '',
                lien_record.pdf_file or ''
            ]
            
            for col, value in enumerate(data_row, 1):
                ws.cell(row=2, column=col, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min((max_length + 2) * 1.2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"lien_record_{lien_record.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            wb.save(response)
            return response
            
        else:
            return JsonResponse({'error': 'Only POST method allowed'}, status=405)
            
    except Exception as e:
        logger.error(f"Error downloading lien Excel: {e}")
        return JsonResponse({'error': str(e)}, status=500)


def download_all_lien_excel(request):
    """Download all lien records as Excel"""
    try:
        # Get all lien records
        lien_records = LienData.objects.all().order_by('-created_at')
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "All Lien Data"
        
        # Add headers with styling
        headers = [
            'ID', 'Direct Party (Debtor)', 'Reverse Party (Claimant)', 'Address', 
            'Zipcode', 'Total Due', 'County', 'Instrument', 'Date Filed',
            'Book', 'Page', 'Description', 'PDF Document URL', 'PDF File', 'Created At'
        ]
        
        # Style headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data rows
        for row_num, record in enumerate(lien_records, 2):
            data_row = [
                record.id,
                record.direct_party_debtor or '',
                record.reverse_party_claimant or '',
                record.address or '',
                record.zipcode or '',
                record.total_due or '',
                record.county or '',
                record.instrument or '',
                record.date_filed or '',
                record.book or '',
                record.page or '',
                record.description or '',
                record.pdf_document_url or '',
                record.pdf_file or '',
                record.created_at.strftime('%Y-%m-%d %H:%M:%S') if record.created_at else ''
            ]
            
            for col, value in enumerate(data_row, 1):
                ws.cell(row=row_num, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2) * 1.2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"all_lien_data_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        logger.error(f"Error downloading all lien Excel: {e}")
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def download_realestate_excel(request):
    """Download real estate records as Excel based on search name"""
    try:
        if request.method == 'POST':
            data = json.loads(request.body)
            search_name = data.get('search_name', '')
            
            # Find real estate records based on search name
            if search_name:
                realestate_records = RealEstateData.objects.filter(search_name__icontains=search_name)
            else:
                realestate_records = RealEstateData.objects.all()
            
            if not realestate_records:
                return JsonResponse({'error': 'No records found'}, status=404)
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Real Estate Data"
            
            # Add headers with styling
            headers = ['Search Name', 'Entity Index', 'Document Index', 'PDF Viewer URL', 'Real Estate PDF URL', 'Created At']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2572a1", end_color="2572a1", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Add data rows
            for row_num, record in enumerate(realestate_records, 2):
                data_row = [
                    record.search_name or 'Not available',
                    record.entity_index or 0,
                    record.doc_index or 0,
                    record.pdf_viewer or 'Not available',
                    record.realestate_pdf or 'Not available',
                    record.created_at.strftime('%Y-%m-%d %H:%M:%S') if record.created_at else 'Not available'
                ]
                
                for col, value in enumerate(data_row, 1):
                    ws.cell(row=row_num, column=col, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min((max_length + 2) * 1.2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"realestate_{search_name or 'all'}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            wb.save(response)
            return response
            
        else:
            return JsonResponse({'error': 'Only POST method allowed'}, status=405)
            
    except Exception as e:
        logger.error(f"Error downloading real estate Excel: {e}")
        return JsonResponse({'error': str(e)}, status=500)


def download_all_realestate_excel(request):
    """Download all real estate records as Excel"""
    try:
        # Get all real estate records
        realestate_records = RealEstateData.objects.all().order_by('-created_at')
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "All Real Estate Data"
        
        # Add headers with styling
        headers = [
            'ID', 'Search Name', 'Entity Index', 'Document Index', 
            'PDF Viewer URL', 'Real Estate PDF URL', 'Created At'
        ]
        
        # Style headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2572a1", end_color="2572a1", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data rows
        for row_num, record in enumerate(realestate_records, 2):
            data_row = [
                record.id,
                record.search_name or '',
                record.entity_index or '',
                record.doc_index or '',
                record.pdf_viewer or '',
                record.realestate_pdf or '',
                record.created_at.strftime('%Y-%m-%d %H:%M:%S') if record.created_at else ''
            ]
            
            for col, value in enumerate(data_row, 1):
                ws.cell(row=row_num, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2) * 1.2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"all_realestate_data_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        logger.error(f"Error downloading all real estate Excel: {e}")
        return JsonResponse({'error': str(e)}, status=500)