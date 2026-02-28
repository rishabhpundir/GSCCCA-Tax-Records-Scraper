from __future__ import annotations

import re
import os
import json
import random
import img2pdf
import traceback
from pathlib import Path
from datetime import datetime
from urllib.parse import urljoin
import html

from openpyxl import Workbook, load_workbook

import pytesseract
import cv2
import numpy as np
import pandas as pd
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from PIL import Image, ImageOps
from rich.console import Console
import playwright.async_api as pw

from dashboard.utils.state import stop_scraper_flag 

try:
    from ocr.realestate_ocr_extractor import extract_re_fields_from_image
except Exception:
    extract_re_fields_from_image = None

load_dotenv()
console = Console()

# ---------- Config -------------------------------------------------------------
HEADLESS = True if os.getenv("HEADLESS", "False").lower() in ("true", "yes") else False
WIDTH, HEIGHT = os.getenv("RES", "1920x1080").split("x")
STATE_FILE = Path("cookies.json")
TAX_EMAIL = os.getenv("GSCCCA_USERNAME")
TAX_PASSWORD = os.getenv("GSCCCA_PASSWORD")
LOCALE = "en-GB"
TIMEZONE = "UTC"
VIEWPORT = {"width": int(WIDTH), "height": int(HEIGHT)}
UA_DICT = {
    "macos": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36",
    "linux": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36",
    "windows": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome"
}

UA = UA_DICT.get(os.getenv("OS_NAME"), "windows")
EXTRA_HEADERS = {"Accept-Language": "en-GB,en-US;q=0.9,en;q=0.8"}

# load Tesseract path for Windows if needed
try:
    if os.name == "nt":  # Windows
        pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
except Exception as e:
    console.print(f"[red]Error setting up Tesseract: {e}[/red]")


# ----------- Directories -----------
BASE_DIR = Path(__file__).parent.absolute() 
BASE_OUTPUT_DIR = os.path.join(BASE_DIR.parent, "output")
os.makedirs(BASE_OUTPUT_DIR, exist_ok=True)

REAL_ESTATE_EXCEL_DIR = os.path.join(BASE_OUTPUT_DIR, "real_estate")
PDF_DIR = os.path.join(REAL_ESTATE_EXCEL_DIR, "documents")

os.makedirs(PDF_DIR, exist_ok=True)
os.makedirs(REAL_ESTATE_EXCEL_DIR, exist_ok=True)


# ---------- Utility Function to find latest Excel file -------------------------
def images_to_pdf(img_paths, pdf_path):
    try:
        valid_images = []
        for path in img_paths:
            try:
                with Image.open(path) as im:
                    im.verify()
                valid_images.append(path)
            except Exception as e:
                console.print(f"[yellow]Skipping invalid image {path}: {e}[/yellow]")

        if not valid_images:
            console.print("[red]No valid images to convert.[/red]")
            return False

        # Convert the list of valid image paths to a single PDF
        for tmp_img in valid_images:
            with Image.open(tmp_img) as im:
                im = ImageOps.exif_transpose(im)
                im = im.rotate(90, expand=True)
                im.save(tmp_img)
                
        pdf_bytes = img2pdf.convert(valid_images)
        with open(pdf_path, "wb") as f:
            f.write(pdf_bytes)

        console.print(f"[green]Successfully created PDF: {pdf_path}[/green]")
        return True

    except Exception as e:
        console.print(f"[red]Failed to create PDF: {e}[/red]")
        return False

# ---------- Scraper Class -----------------------------------------------------
class RealEstateIndexScraper:
    def __init__(self) -> None:
        try:
            self.playwright = None
            self.browser = None
            self.page = None
            self.email = TAX_EMAIL
            self.password = TAX_PASSWORD
            self.homepage_url = "https://www.gsccca.org/"
            self.login_url = "https://apps.gsccca.org/login.asp"
            self.realestate_search_url = "https://search.gsccca.org/RealEstate/namesearch.asp"
            self.results = []
            self.form_data = {}
            
            # Use global constants defined above
            self.pdf_dir = PDF_DIR
            self.excel_output_dir = REAL_ESTATE_EXCEL_DIR
            console.print(f"[green]Real Estate Data Output directory --> {self.excel_output_dir}[/green]")
        except Exception as e:
            console.print(f"[red]Error initializing RealEstateIndexScraper: {e}[/red]")
            raise


    def time_sleep(self, a: int = 2500, b: int = 5000) -> int:
        return random.uniform(a, b)
    

    async def stop_check(self):
        """ Global stop flag to immediately exit scraping if invoked by user. """
        if stop_scraper_flag['realestate']:
            console.print("[yellow]Lien Index Scraper received immediate stop signal. Exiting...[/yellow]")
            if self.browser:
                await self.browser.close()
            if self.playwright:
                await self.playwright.stop()
            raise pw.Error("STOP_REQUESTED")


    # ----------------- OCR / FIELD EXTRACTION ----------------- #
    RE_SKIP_WORDS = ("CANCELLATION", "CANCELLED", "FORECLOSURE", "FORECLOSED")

    def _contains_skip_words(self, text: str) -> bool:
        up = (text or "").upper()
        return any(w in up for w in self.RE_SKIP_WORDS)

    def _first_match(self, patterns, text: str) -> str:
        for pat in patterns:
            m = re.search(pat, text, re.I | re.S)
            if m:
                return (m.group(1) or "").strip()
        return ""

    def _extract_money(self, text: str) -> str:
        # prefer bigger amounts; return numeric string without $ and commas
        monies = re.findall(r"\$\s*([\d,]+(?:\.\d{1,2})?)", text or "")
        if not monies:
            return ""
        def to_float(s):
            try:
                return float(s.replace(",", ""))
            except Exception:
                return 0.0
        best = max(monies, key=to_float)
        return best.replace(",", "").strip()

    def _extract_dates(self, text: str):
        # returns (mortgage_date, assignment_date) best-effort as raw strings
        t = text or ""

        # mortgage/original date cues
        mortgage_patterns = [
            r"made\s+this\s+\d{1,2}(?:st|nd|rd|th)?\s+day\s+of\s+([A-Za-z]+\s+\d{1,2},\s*\d{4})",
            r"given\s+on\s+([A-Za-z]+\s+\d{1,2},\s*\d{4})",
            r"dated\s+([A-Za-z]+\s+\d{1,2},\s*\d{4})",
            r"dated\s+(\d{1,2}/\d{1,2}/\d{2,4})",
        ]
        mortgage_date = self._first_match(mortgage_patterns, t)

        # assignment date cues (often "Filed and Recorded ..." header or "this __ day of ...")
        assign_patterns = [
            r"Filed\s+and\s+Recorded\s+([A-Za-z]+\s+\d{1,2},\s*\d{4})",
            r"Filed\s+and\s+Recorded\s+.*?(\d{1,2}/\d{1,2}/\d{2,4})",
            r"this\s+\d{1,2}(?:st|nd|rd|th)?\s+day\s+of\s+([A-Za-z]+\s*,\s*\d{4})",
            r"this\s+\d{1,2}(?:st|nd|rd|th)?\s+day\s+of\s+([A-Za-z]+\s+\d{4})",
        ]
        assignment_date = self._first_match(assign_patterns, t)

        return mortgage_date, assignment_date

    def _extract_name(self, text: str) -> str:
        t = text or ""
        patterns = [
            r"grantor\s+is\s+([A-Z][A-Z\s\.,'-]{3,80})",
            r"Borrower\)?\s*[:\-]?\s*([A-Z][A-Z\s\.,'-]{3,80})",
            r"\(Borrower\)\s*([A-Z][A-Z\s\.,'-]{3,80})",
        ]
        name = self._first_match(patterns, t)
        # cleanup trailing label words
        name = re.split(r"\bwhose\b|\bwho\b|\baddress\b|\bherein\b", name, flags=re.I)[0].strip(" ,;")
        return name

    def _extract_original_lender(self, text: str) -> str:
        t = text or ""
        patterns = [
            r"in\s+favor\s+of\s*\(Lender\)\s*([A-Z][A-Z0-9\s,&\.-]{3,80})",
            r"\(Lender\)\s*[:\-]?\s*([A-Z][A-Z0-9\s,&\.-]{3,80})",
            r"Lender\s*[:\-]\s*([A-Z][A-Z0-9\s,&\.-]{3,80})",
        ]
        lender = self._first_match(patterns, t)
        lender = lender.splitlines()[0].strip(" ,;")
        return lender

    def _extract_property_address(self, text: str, ocr_addresses=None) -> str:
        t = text or ""
        patterns = [
            r"Property\s+Address\W*([\s\S]{0,120}?\bGA\b\s*\d{5}(?:-\d{4})?)",
            r"located\s+at\s+([\s\S]{0,120}?\bGA\b\s*\d{5}(?:-\d{4})?)",
            r"whose\s+address\s+is\s+([\s\S]{0,120}?\bGA\b\s*\d{5}(?:-\d{4})?)",
        ]
        addr = self._first_match(patterns, t)
        addr = re.sub(r"\s+", " ", addr).strip(" ,;:-")
        if addr:
            return addr

        # fallback: pick first GA ZIP address block from OCR extractor
        if ocr_addresses:
            for a in ocr_addresses:
                if isinstance(a, str) and re.search(r"\bGA\b\s*\d{5}", a, re.I):
                    return re.sub(r"\s+", " ", a).strip()
        return ""

    def _extract_re_fields_from_ocr(self, raw_text: str, ocr_json: dict | None) -> dict:
        # best-effort field extraction for Excel columns
        mortgage_date, assignment_date = self._extract_dates(raw_text)

        ocr_addresses = []
        if isinstance(ocr_json, dict):
            ocr_addresses = ocr_json.get("addresses") or []

        # Mortgage amount: prefer OCR JSON top amount if available
        mortgage_amount = ""
        if isinstance(ocr_json, dict):
            top = (ocr_json.get("amounts") or {}).get("top_by_score") or []
            if top and isinstance(top[0], dict):
                mortgage_amount = str(top[0].get("numeric") or "").strip()
        if not mortgage_amount:
            mortgage_amount = self._extract_money(raw_text)

        return {
            "Name": self._extract_name(raw_text),
            "Mortgage Date (original)": mortgage_date,
            "Assignment Date": assignment_date,
            "Original Lender": self._extract_original_lender(raw_text),
            "Mortgage Amount": mortgage_amount,
            "Property Address": self._extract_property_address(raw_text, ocr_addresses=ocr_addresses),
        }


    async def dump_cookies(self, out_file="cookies.json"):
        """Save cookies + storage ONLY for login check."""
        try:
            state = await self.page.context.storage_state()
            Path(out_file).write_text(json.dumps(state, indent=2))
            print(f"Saved login state to --> {out_file}")
        except Exception as e:
            console.print(f"[red]Failed to dump cookies: {e}[/red]")


    async def already_logged_in(self) -> bool:
        """Check if user is already logged in."""
        try:
            await self.page.wait_for_timeout(self.time_sleep())
            all_text = await self.page.evaluate("document.body.innerText")
            if "logout" in all_text.lower():
                return True
            return False

        except Exception as e:
            print(f"[already_logged_in ERROR] {e}")
            return False

    
    async def check_session(self):
        """Check on homepage if user is logged in."""
        try:
            if await self.already_logged_in():
                console.print("[green]Logged in session detected!![/green]")
                return True
            else:
                console.print("[red]Session not found![/red]")
                return False
        except Exception as e:
            console.print(f"[red]Error in check_session: {e}[/red]")
            return False


    async def check_and_handle_announcement(self):
        """Check if announcement page loaded; if yes, redirect to realestate_search_url."""
        try:
            current_url = self.page.url
            if "Announcement" in current_url:
                await self.page.select_option("#Options", "dismiss")
                await self.page.wait_for_timeout(1000)
                await self.page.click("input[name='Continue']")
                print("Announcement page detected. Turning off...")
        except Exception as e:
            console.print(f"[red]Error handling announcement: {e}[/red]")
            

    async def login(self):
        """Perform login and save cookies."""
        try:
            await self.page.goto(self.login_url, wait_until="domcontentloaded", timeout=60000)
            await self.page.wait_for_timeout(self.time_sleep())
            await self.check_and_handle_announcement()

            await self.page.fill("input[name='txtUserID']", self.email)
            await self.page.wait_for_timeout(self.time_sleep()) 
            await self.page.fill("input[name='txtPassword']", self.password)
            await self.page.wait_for_timeout(self.time_sleep())
            checkbox = await self.page.query_selector("input[type='checkbox'][name='permanent']")
            await self.page.wait_for_timeout(self.time_sleep(a=1000, b=1200))
            
            if checkbox:
                is_checked = await checkbox.is_checked()
                if not is_checked:
                    await checkbox.click()
            else:
                print("[LOGIN] Checkbox not found on the page.")

            try:
                await self.page.click("img[name='logon']")
            except Exception as e:
                print(f"[LOGIN] Login button Click failed: {e}, using JS submit...")
                await self.page.evaluate("document.forms['frmLogin'].submit()")

            await self.page.wait_for_load_state("networkidle", timeout=60000)
            await self.page.wait_for_timeout(self.time_sleep())
            await self.check_and_handle_announcement()
            
            await self.page.goto(self.realestate_search_url, wait_until="domcontentloaded", timeout=60000)
            await self.page.wait_for_timeout(self.time_sleep())
            await self.check_and_handle_announcement()
            
            if await self.already_logged_in():
                console.print("[bold green]Login successful![/bold green]")
                await self.dump_cookies()
                return True
            else:
                print("[LOGIN] Login failed!")
                return False
        except Exception as e:
            console.print(f"[red]Error during login: {e}[/red]")
            return False
    
    
    async def start_realestate_search(self):
        """Fill the real estate form using the provided parameters."""
        try:
            print("Executing Real Estate Term search...")
            await self.page.goto(self.realestate_search_url, wait_until="domcontentloaded", timeout=60000)
            await self.page.wait_for_timeout(self.time_sleep())
            await self.check_and_handle_announcement()
        except Exception as e:
            console.print(f"[red]Error in start_realestate_search: {e}[/red]")
            
        try:
            await self.page.wait_for_selector("input[name='txtSearchName']")
            await self.page.wait_for_timeout(self.time_sleep(a=250, b=500))
            await self.page.select_option("select[name='txtPartyType']", self.form_data.get("txtPartyType", "2"))
            await self.page.wait_for_timeout(self.time_sleep(a=250, b=500))
            await self.page.select_option("select[name='txtInstrCode']", self.form_data.get("txtInstrCode", "ALL"))
            await self.page.wait_for_timeout(self.time_sleep(a=250, b=500))
            await self.page.select_option("select[name='intCountyID']", self.form_data.get("intCountyID", "-1"))
            await self.page.wait_for_timeout(self.time_sleep(a=250, b=500))
            
            include_val = self.form_data.get("bolInclude", "0")
            checkbox_selector = f"input[name='bolInclude'][value='{include_val}']"
            if await self.page.query_selector(checkbox_selector):
                await self.page.check(checkbox_selector)
            
            await self.page.fill("input[name='txtSearchName']", self.form_data.get("txtSearchName", ""))
            await self.page.wait_for_timeout(self.time_sleep(a=250, b=500))
            await self.page.fill("input[name='txtFromDate']", self.form_data.get("txtFromDate", ""))
            await self.page.wait_for_timeout(self.time_sleep(a=250, b=500))
            await self.page.fill("input[name='txtToDate']", self.form_data.get("txtToDate", ""))
            await self.page.wait_for_timeout(self.time_sleep(a=250, b=500))
            
            await self.page.select_option("select[name='MaxRows']", self.form_data.get("MaxRows", "100"))
            await self.page.wait_for_timeout(self.time_sleep(a=250, b=500))
            await self.page.select_option("select[name='TableType']", self.form_data.get("TableType", "1"))
            
            await self.page.wait_for_timeout(self.time_sleep())
            await self.page.click("#btnSubmit")
        except Exception as e:
            console.print(f"[red]Error filling real estate form: {e}[/red]\n{traceback.format_exc()}")
            raise


    async def get_search_results(self):
        """Step: Extract ALL document URLs first, then save them to CSV for processing."""
        try:
            print(f"Loading Real Estate Search results...")
            await self.page.wait_for_timeout(self.time_sleep(a=4000, b=5000))

            search_name = (self.form_data.get("txtSearchName") or "").strip()
            safe_search = re.sub(r"[^a-zA-Z0-9]+", "_", search_name).strip("_") or "search"
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")

            # session artifacts
            self.csv_path = os.path.join(self.excel_output_dir, f"{safe_search}_realestate_urls_{ts}.csv")
            self.excel_path = os.path.join(self.excel_output_dir, f"realestate_index_{safe_search}_{ts}.xlsx")

            radios = await self.page.query_selector_all("input[name='rdoEntityName']")
            print("*" * 50)
            console.print(f"[cyan]Found {len(radios)} entity result(s) for '{search_name}'[/cyan]")

            results_df = pd.DataFrame(columns=["url", "status", "search_name", "entity_index", "doc_index"])

            for entity_idx in range(1, len(radios) + 1):
                # if len(results_df) >= 10:
                #     break
                await self.stop_check()
                try:
                    # Refresh the radios list to avoid staleness
                    current_radios = await self.page.query_selector_all("input[name='rdoEntityName']")
                    radio = current_radios[entity_idx - 1] if entity_idx - 1 < len(current_radios) else None
                    if not radio:
                        console.print(f"[yellow]Radio button for entity {entity_idx} not found[/yellow]")
                        continue

                    await radio.scroll_into_view_if_needed()
                    await radio.wait_for_element_state("visible", timeout=10000)
                    await radio.click()

                    # Display Details
                    await self.page.click("#btnDisplayDetails")
                    await self.page.wait_for_load_state("domcontentloaded", timeout=20000)
                    await self.page.wait_for_timeout(self.time_sleep())
                    await self.check_and_handle_announcement()

                    # ---- Extract GE/GR document links for this entity (NO navigation here) ----
                    await self.page.wait_for_timeout(self.time_sleep(a=500, b=1200))

                    hrefs = await self.page.eval_on_selector_all(
                        "a[href*='final.asp']",
                        "els => els.map(e => e.getAttribute('href'))"
                    )

                    base = "https://search.gsccca.org/RealEstate/"
                    urls = []
                    js_pat = re.compile(r"fnSubmitThisForm\('([^']+)'\)")

                    for h in hrefs or []:
                        if not h:
                            continue
                        if "fnSubmitThisForm" in h:
                            m = js_pat.search(h)
                            if not m:
                                continue
                            rel = html.unescape(m.group(1))
                            urls.append(urljoin(base, rel))
                        else:
                            urls.append(urljoin(base, h))

                    urls = list(dict.fromkeys(urls))  # de-dupe, preserve order

                    console.print(f"[green]Entity {entity_idx}: extracted {len(urls)} document url(s)[/green]")

                    if urls:
                        tmp_df = pd.DataFrame({
                            "url": urls,
                            "status": [""] * len(urls),
                            "search_name": [search_name] * len(urls),
                            "entity_index": [entity_idx] * len(urls),
                            "doc_index": list(range(1, len(urls) + 1)),
                        })
                        results_df = pd.concat([results_df, tmp_df], ignore_index=True)

                    # Go back to entity selection page
                    back_ok = False
                    try:
                        await self.page.go_back()
                        await self.page.wait_for_load_state("domcontentloaded", timeout=20000)
                        await self.page.wait_for_timeout(self.time_sleep())
                        back_ok = True
                    except Exception:
                        back_ok = False

                    if not back_ok:
                        # Recovery: go back to search page and re-run the search
                        console.print("[yellow]Recovery: returning to Real Estate search page and re-searching...[/yellow]")
                        await self.start_realestate_search()
                        await self.page.wait_for_timeout(self.time_sleep())
                        await self.page.wait_for_selector("input[name='rdoEntityName']", timeout=30000)

                except Exception as e:
                    console.print(f"[red]Error extracting URLs for entity {entity_idx}: {e}[/red]")
                    continue

            # save urls list
            results_df.drop_duplicates(subset=["url"], inplace=True)
            results_df.reset_index(drop=True, inplace=True)
            results_df.to_csv(self.csv_path, index=False)
            console.print(f"[green]Success -> Search results' URLs saved to CSV at {self.csv_path}[/green]")

        except Exception as e:
            console.print(f"[red]Step get_search_results error: {e}[/red]\n{traceback.format_exc()}")



    async def process_result_urls(self):
        """Process all extracted document URLs (CSV) and save results incrementally."""
        if not getattr(self, "csv_path", None) or not os.path.exists(self.csv_path):
            console.print(f"[red][ERROR] CSV URL list not found: {getattr(self, 'csv_path', None)}[/red]")
            return

        df_urls = pd.read_csv(self.csv_path)
        if df_urls.empty:
            console.print(f"[red][ERROR] No URLs found at: {self.csv_path}[/red]")
            return

        console.print(f"[cyan]Initiating Real Estate data extraction... Total URLs: {len(df_urls)}[/cyan]")

        for idx, row in df_urls.iterrows():
            if str(row.get("status", "")).strip().lower() == "done":
                continue

            await self.stop_check()

            url = str(row.get("url", "")).strip()
            search_name = str(row.get("search_name", "")).strip()
            entity_idx = int(row.get("entity_index", 0) or 0)
            doc_idx = int(row.get("doc_index", 0) or 0)

            if not url:
                df_urls.at[idx, "status"] = "Done"
                df_urls.to_csv(self.csv_path, index=False)
                continue

            print("-" * 50)
            console.print(f"[green]{idx + 1}. Processing Entity {entity_idx}, Doc {doc_idx}[/green]")
            console.print(f"[blue]URL: {url}[/blue]")

            try:
                await self.page.goto(url, wait_until="domcontentloaded", timeout=60000)
                await self.page.wait_for_timeout(self.time_sleep())
                await self.check_and_handle_announcement()

                data = await self.parse_realestate_data(
                    search_name=search_name,
                    entity_idx=entity_idx,
                    doc_idx=doc_idx,
                    source_url=url,
                )

                if data is None:
                    console.print(f"[yellow]Skipped (cancelled/foreclosed) -> {row['url']}[/yellow]")
                elif data:
                    self.results.append(data)
                    await self._append_result_to_excel(data)
                    console.print(f"[cyan]Saved record for Entity {entity_idx}, Doc {doc_idx}[/cyan]")
                else:
                    console.print(f"[yellow]No data extracted for Entity {entity_idx}, Doc {doc_idx}[/yellow]")

                df_urls.at[idx, "status"] = "Done"
                df_urls.to_csv(self.csv_path, index=False)

            except Exception as e:
                console.print(f"[red]Error processing URL {url}: {e}[/red]\n{traceback.format_exc()}")
                # keep it un-done for resume
                continue


    async def parse_realestate_data(self, search_name: str, entity_idx: int, doc_idx: int, source_url: str):
        """Parse one Real Estate document detail page and generate a PDF from the HTML5 viewer."""
        await self.stop_check()

        data = {
            "Search Name": search_name,
            "Entity Index": entity_idx,
            "Doc Index": doc_idx,
            "Source URL": source_url,
        }

        popup = None
        try:
            html_text = await self.page.content()
            soup = BeautifulSoup(html_text, "html.parser")

            # ---------- PDF Viewer URL Extraction ----------
            viewer_script = soup.find("script", string=lambda t: t and "ViewImage" in t)
            if not viewer_script or not viewer_script.string:
                data["PDF Viewer URL"] = "ADD_TAG"
                return data

            script_text = viewer_script.string

            # NOTE: Keep existing tags/vars where possible; placeholders if missing
            reid_match = re.search(r"var iREID\s*=\s*(\d+);", script_text)  # RealEstate id
            if not reid_match:
                data["PDF Viewer URL"] = "ADD_TAG"
                return data

            reid = reid_match.group(1)

            # These vars exist in your current scraper; if any missing, use placeholder.
            def _pick(pattern: str, placeholder: str = "ADD_TAG") -> str:
                m = re.search(pattern, script_text)
                return m.group(1) if m else placeholder

            county = _pick(r"var county\s*=\s*\"(\d+)\"")
            book = _pick(r"var book\s*=\s*\"(\d+)\"")
            page_num = _pick(r"var page\s*=\s*\"(\d+)\"")
            userid = _pick(r"var user\s*=\s*(\d+)")
            appid = _pick(r"var appid\s*=\s*(\d+)")
            data["Book"] = book
            data["Page"] = page_num

            viewer_url = (
                "https://search.gsccca.org/Imaging/HTML5Viewer.aspx?"  # viewer base
                f"id={reid}&key1={book}&key2={page_num}&county={county}&userid={userid}&appid={appid}"
            )
            data["PDF Viewer URL"] = viewer_url

            popup = await self.page.context.new_page()
            await popup.goto(viewer_url, wait_until="domcontentloaded", timeout=60000)
            await popup.wait_for_timeout(6000)

            # --- Collect thumbnails/pages ---
            thumb_links = await popup.query_selector_all("a[id*='lvThumbnails_lnkThumbnail']")
            pages_count = len(thumb_links)
            data["Pages"] = pages_count

            if pages_count == 0:
                console.print("[yellow]No thumbnails found in viewer[/yellow]")
                data["Real Estate PDF"] = ""
                return data

            screenshot_paths = []

            # Try to set fit window once (best-effort)
            try:
                await popup.wait_for_selector("td.vtm_zoomSelectCell select", timeout=10000)
                await popup.select_option("td.vtm_zoomSelectCell select", "fitwindow")
                await popup.wait_for_timeout(1500)
            except Exception:
                pass

            # Rotate right once (best-effort, matches lien flow)
            try:
                await popup.locator('img[title="Rotate Right"]').click()
                await popup.wait_for_timeout(1000)
            except Exception:
                pass

            await popup.wait_for_selector("div.vtm_imageClipper canvas", timeout=20000, state="attached")
            canvas = await popup.query_selector("div.vtm_imageClipper canvas")

            safe_base = f"RE_Entity_{entity_idx}_Doc_{doc_idx}"
            # Best effort: read header for book/page
            try:
                header_text = await popup.inner_text("#lblHeader")
                m = re.search(r"Book\s+(\d+)\s+Page\s+(\d+)", header_text)
                if m:
                    safe_base = f"RE_Book_{m.group(1)}_Page_{m.group(2)}_Entity_{entity_idx}_Doc_{doc_idx}"
            except Exception:
                pass

            await self.stop_check()
            try:
                await popup.wait_for_timeout(self.time_sleep())
                if not canvas:
                    console.print("[yellow]Canvas not found for screenshot[/yellow]")

                screenshot_path = Path(os.path.join(self.pdf_dir, f"{safe_base}_Page_{page_num}.png"))
                await canvas.screenshot(path=str(screenshot_path), timeout=30000)
                screenshot_paths.append(screenshot_path)
            except Exception as e:
                console.print(f"[red]Error processing thumbnail: {e}[/red]")

            if not screenshot_paths:
                data["Real Estate PDF"] = ""
                return data

            pdf_path = Path(os.path.join(self.pdf_dir, f"{safe_base}.pdf"))
            images_to_pdf(screenshot_paths, pdf_path)

            # cleanup pngs
            # for p in screenshot_paths:
            #     try:
            #         os.remove(p)
            #     except FileNotFoundError:
            #         pass

            data["Real Estate PDF"] = str(pdf_path)

            # Optional OCR (placeholder tag if not needed / not implemented here)
            # ---------- OCR Extraction (like lien scraper) ----------
            ocr_raw_parts = []
            ocr_json = None
            try:
                ocr_raw_text = "\n\n".join(ocr_raw_parts).strip()
                data["OCR Raw Text"] = ocr_raw_text if ocr_raw_text else ""

                # Skip cancelled/foreclosed docs
                if self._contains_skip_words(data.get("OCR Raw Text", "")):
                    data["SKIP_REASON"] = "CANCELLED/FORECLOSED"
                    return None

                # Skip cancelled/foreclosed docs (using OCR engine dict result)
                if extract_re_fields_from_image and screenshot_paths:
                    fields = extract_re_fields_from_image(
                        img_path=str(screenshot_paths[0]),        # or loop all pages if you want later
                        use_paddle=False,       # or hardcode True/False
                        cache_dir=".re_ocr_cache",                # optional cache
                        debug=False
                    )

                    # if the OCR layer says skip — skip it
                    if fields.get("SKIP_REASON"):
                        data["SKIP_REASON"] = fields["SKIP_REASON"]
                        return None

                    data.update(fields)
                else:
                    # fallback to your existing logic
                    fields = self._extract_re_fields_from_ocr(raw_text=data.get("OCR Raw Text",""), ocr_json=ocr_json)
                    data.update(fields)

            except Exception as e:
                console.print(f"[yellow]OCR extraction failed: {e}[/yellow]")
                data["OCR Raw Text"] = data.get("OCR Raw Text","") or ""

            return data

        except Exception as e:
            console.print(f"[bold red]Fatal error in parse_realestate_data: {e}[/bold red]\n{traceback.format_exc()}")
            return data
        finally:
            try:
                if popup and popup != self.page:
                    await popup.close()
            except Exception:
                pass


    async def parse_documents(self, search_name: str, entity_idx: int):
        """DEPRECATED: kept for backward-compat; use get_search_results() + process_result_urls()."""
        console.print("[yellow]parse_documents() is deprecated. URLs are now extracted first, then processed from CSV.[/yellow]")
        return


    def _excel_safe(self, v):
        if v is None:
            return ""
        if isinstance(v, str):
            return v.replace("\x00", "").strip()
        return v

    async def _append_result_to_excel(self, data: dict):
        """Append ONE row to an Excel file (atomic write), similar to lien scraper."""
        if not getattr(self, "excel_path", None):
            # fallback
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.excel_path = os.path.join(self.excel_output_dir, f"realestate_index_{ts}.xlsx")

        headers = [
            "Name",
            "Mortgage Date (original)",
            "Assignment Date",
            "Original Lender",
            "Mortgage Amount",
            "Property Address",
            "Search Name",
            "Entity Index",
            "Doc Index",
            "Book",
            "Page",
            "Pages",
            "PDF Viewer URL",
            "Source URL",
            "View PDF",
        ]

        row_vals = []
        for h in headers:
            if h == "View PDF":
                pdf_path = str(data.get("Real Estate PDF", "") or "")
                if pdf_path:
                    pdf_name = os.path.basename(pdf_path)
                    view_path = pdf_path.replace(os.sep, "/")
                    row_vals.append(f'=HYPERLINK("file:///{view_path}", "{pdf_name}")')
                else:
                    row_vals.append("")
            else:
                row_vals.append(self._excel_safe(data.get(h, "")))

        if os.path.exists(self.excel_path):
            wb = load_workbook(self.excel_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Real Estate Data"
            ws.append(headers)

        ws.append(row_vals)

        tmp_path = self.excel_path + ".tmp"
        wb.save(tmp_path)
        os.replace(tmp_path, self.excel_path)
        console.print(f"[bold green]Saved record to --> {self.excel_path}[/bold green]")


    def save_results_to_excel(self, filename_prefix="realestate_index"):
        """Legacy full-save. If incremental Excel was used, this simply returns the session excel path."""
        if getattr(self, "excel_path", None) and os.path.exists(self.excel_path):
            console.print(f"[green]Session Excel already exists -> {self.excel_path}[/green]")
            return self.excel_path

        if not self.results:
            console.print("[red]No results to save[/red]")
            return None

        try:
            df = pd.DataFrame(self.results)
            df.drop_duplicates(subset=["Search Name", "Real Estate PDF"], inplace=True)
            df.reset_index(drop=True, inplace=True)

            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            final_filename = f"{filename_prefix}_{ts}.xlsx"
            final_path = os.path.join(self.excel_output_dir, final_filename)

            with pd.ExcelWriter(final_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Real Estate Data', index=False)

            console.print(f"[green]Real Estate Excel saved -> {final_path}[/green]")
            return final_path

        except Exception as e:
            console.print(f"[red]Failed to save Excel: {e}[/red]")
            traceback.print_exc()
            return None


    async def scrape(self, formdata: dict):
        try:
            self.form_data = formdata
            self.playwright = await pw.async_playwright().start()
            self.browser = await self.playwright.chromium.launch(
                headless=HEADLESS,
                channel="chrome",
                args=[
                    "--disable-blink-features=AutomationControlled",
                    "--start-maximized",
                    "--no-proxy-server",
                ]
            )

            print("Starting Real Estate Index Scraper...")
            print(f"Screen Resolution set to --> {WIDTH}x{HEIGHT}")
            context = await self.browser.new_context(
                storage_state=STATE_FILE if STATE_FILE.exists() else None,
                user_agent=UA,
                locale=LOCALE,
                timezone_id=TIMEZONE,
                viewport=VIEWPORT,
                device_scale_factor=1,
                extra_http_headers=EXTRA_HEADERS,
                bypass_csp=True,
                ignore_https_errors=False, 
            )
            self.page = await context.new_page()
            self.context = context

            # login if needed
            await self.page.goto("https://google.com", wait_until="domcontentloaded")
            await self.page.wait_for_timeout(self.time_sleep())
            if STATE_FILE.exists():
                # await context.add_cookies(json.loads(Path(STATE_FILE).read_text())["cookies"])
                await self.page.goto(self.homepage_url, wait_until="domcontentloaded", timeout=60000)
                await self.check_and_handle_announcement()
            else:
                await self.page.goto(self.login_url, wait_until="domcontentloaded")
                await self.check_and_handle_announcement()
                if not await self.check_session():
                    print("Attempting fresh login...")
                    await self.login()
                    await self.page.wait_for_timeout(self.time_sleep())
                        
            await self.stop_check()
            if not await self.check_session():
                console.print("[yellow]Session invalid... logging in again...[/yellow]")
                await self.login()
                await self.page.wait_for_timeout(self.time_sleep())

            # Start real estate index search
            await self.start_realestate_search()

            await self.get_search_results()
            await self.process_result_urls()

        except Exception as e:
            console.print(f"[red]Error in scrape method: {e}[/red]")
            traceback.print_exc()
        finally:
            if self.browser:
                await self.browser.close()
            if self.playwright:
                await self.playwright.stop()

