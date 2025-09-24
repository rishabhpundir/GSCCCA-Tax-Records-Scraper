# dashboard/views.py
from django.shortcuts import render
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import threading
import json
import pandas as pd
import os
from pathlib import Path
from datetime import datetime
import asyncio
from dashboard.models import LienData, RealEstateData
import traceback
import logging
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import json
from django.utils import timezone

# ------------------ LOGGER SETUP -------------------
logger = logging.getLogger(__name__)

# Add this base directory path
BASE_DIR = Path(__file__).resolve().parent.parent
SCRAPERS_DIR = BASE_DIR / "scrapers"
OUTPUT_DIR = SCRAPERS_DIR / "Output"
DOWNLOADS_DIR = SCRAPERS_DIR / "downloads"



def my_view(request):
    logger.info("User opened dashboard view")
    logger.debug("This is a debug message for developers")
    logger.error("Something went wrong!")
    
def dashboard(request):
    lien_data = LienData.objects.all().order_by('-created_at')
    realestate_data = RealEstateData.objects.all().order_by('-created_at')
    
    return render(request, 'dashboard.html', {
        'lien_data': lien_data,
        'realestate_data': realestate_data
    })

@csrf_exempt
def start_scraper(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        scraper_type = data.get('scraper_type')
        
        if scraper_type == 'lien':
            thread = threading.Thread(target=run_lien_scraper_and_save)
            thread.start()
            return JsonResponse({'status': 'Lien scraper started'})
        
        elif scraper_type == 'realestate':
            thread = threading.Thread(target=run_realestate_scraper_and_save)
            thread.start()
            return JsonResponse({'status': 'Real estate scraper started'})
    
    return JsonResponse({'error': 'Invalid request'}, status=400)

def get_latest_data(request):
    data_type = request.GET.get('type', 'lien')
    
    if data_type == 'lien':
        data = list(LienData.objects.all().order_by('-created_at').values())
    else:
        data = list(RealEstateData.objects.all().order_by('-created_at').values())
    
    return JsonResponse({'data': data})


def run_lien_scraper_and_save():
    """Run lien scraper and save results to database"""
    try:
        logger.info("Starting lien scraper...")
        
        # Import the scraper module
        from scrapers.lien_index_scraper import GSCCCAScraper
        
        # Run the lien scraper
        scraper = GSCCCAScraper()
        asyncio.run(scraper.run())
        
        # Find the latest Excel file - check the correct Output directory
        latest_file = find_latest_excel_file(OUTPUT_DIR, "LienResults")
        
        if not latest_file:
            # Check current scraper directory as fallback
            latest_file = find_latest_excel_file(SCRAPERS_DIR, "LienResults")
            
        if latest_file:
            logger.info(f"Found lien Excel file: {latest_file}")
            
            # Read and save to database
            df = pd.read_excel(latest_file)
            logger.debug(f"Excel file columns: {list(df.columns)}")
            logger.info(f"Number of rows: {len(df)}")
            
            saved_count = 0
            for index, row in df.iterrows():
                try:
                    # Helper function to safely extract and format data
                    def get_field_value(field_name, default='', max_length=None):
                        value = row.get(field_name, default)
                        if pd.isna(value):
                            value = default
                        
                        # Convert to string for consistency
                        value = str(value) if value is not None else default
                        
                        # Apply length limit if specified
                        if max_length is not None and value:
                            return value[:max_length]
                        return value
                    
                    # Create or update record
                    obj, created = LienData.objects.update_or_create(
                        direct_party_debtor=get_field_value('Direct Party (Debtor)', max_length=255),
                        reverse_party_claimant=get_field_value('Reverse Party (Claimant)', max_length=255),
                        book=get_field_value('Book', max_length=50),
                        page=get_field_value('Page', max_length=50),
                        defaults={
                            'address': get_field_value('Address'),
                            'zipcode': get_field_value('Zipcode', max_length=10),
                            'total_due': get_field_value('Total Due', max_length=50),
                            'county': get_field_value('County', max_length=100),
                            'instrument': get_field_value('Instrument', max_length=50),
                            'date_filed': get_field_value('Date Filed', max_length=50),
                            'description': get_field_value('Description'),
                            'pdf_document_url': get_field_value('PDF Document URL'),
                            'pdf_file': get_field_value('PDF', max_length=255),
                        }
                    )
                    
                    if created:
                        saved_count += 1
                        logger.debug(f"Saved record {index + 1}: {get_field_value('Direct Party (Debtor)')}")
                        
                except Exception as e:
                    logger.error(f"Error saving row {index + 1}: {e}")
                    logger.debug(f"Problematic row data: {dict(row)}")
                    traceback.print_exc()
                    
            logger.info(f"Successfully saved {saved_count} out of {len(df)} lien records to database")
        else:
            logger.warning("No lien Excel file found")
            # Check what files exist in the correct locations
            logger.info("Files in Output directory: %s", list(OUTPUT_DIR.glob("*")) if OUTPUT_DIR.exists() else "Output directory doesn't exist")
            logger.info("Files in downloads directory: %s", list(DOWNLOADS_DIR.glob("*")) if DOWNLOADS_DIR.exists() else "Downloads directory doesn't exist")
            logger.info("Files in scrapers directory: %s", list(SCRAPERS_DIR.glob("LienResults*")))

    except Exception as e:
        logger.error(f"Error running lien scraper: {e}")
        traceback.print_exc()

def run_realestate_scraper_and_save():
    """Run real estate scraper and save results to database"""
    try:
        logger.info("Starting real estate scraper...")
        
        # Import the scraper module
        from scrapers.realestate_index_scraper import RealestateIndexScraper
        
        # Run the real estate scraper
        scraper = RealestateIndexScraper()
        asyncio.run(scraper.run())

        # Agar results hain, to pehle unhe database me save karo
        if hasattr(scraper, 'results') and scraper.results:
            logger.info(f"Found {len(scraper.results)} results in memory, saving to database first.")
            
            saved_count = 0
            for result in scraper.results:
                try:
                    # 'search_name' ko 'Search Name' se map kiya
                    search_name = result.get('Search Name', '')
                    entity_index = int(result.get('Entity Index', 0) or 0)
                    doc_index = int(result.get('Doc Index', 0) or 0)
                    pdf_viewer_url = result.get('PDF Viewer URL', '')
                    realestate_pdf_path = result.get('Real Estate PDF', '')

                    # Django ORM se database mein data save karo
                    obj, created = RealEstateData.objects.update_or_create(
                        search_name=search_name[:255],
                        entity_index=entity_index,
                        doc_index=doc_index,
                        defaults={
                            'pdf_viewer': pdf_viewer_url,
                            'realestate_pdf': realestate_pdf_path,
                        }
                    )
                    
                    if created:
                        saved_count += 1
                        logger.debug(f"Saved new real estate record: {search_name}")
                        
                except Exception as e:
                    logger.error(f"Error saving real estate result to DB: {e}")
                    logger.debug(f"Problematic result: {result}")
                    traceback.print_exc()

            logger.info(f"Successfully saved {saved_count} real estate records to database.")

            # Ab, database se data nikal kar Excel file mein save karo
            excel_path = scraper.save_results_to_excel()
            if excel_path:
                logger.info(f"✓ Real Estate data successfully saved to Excel at: {excel_path}")
            else:
                logger.error("✗ Failed to save Excel file from scraper results.")
        
        else:
            logger.warning("No real estate results found in scraper, nothing to save.")
            
    except Exception as e:
        logger.error(f"Error running real estate scraper: {e}")
        traceback.print_exc()

        # Agar results hain, to pehle unhe database me save karo
        if hasattr(scraper, 'results') and scraper.results:
            logger.info(f"Found {len(scraper.results)} results in memory, saving to database first.")

            saved_count = 0
            for result in scraper.results:
                try:
                    # 'search_name' ko 'Search Name' se map kiya
                    search_name = result.get('Search Name', '')
                    entity_index = int(result.get('Entity Index', 0) or 0)
                    doc_index = int(result.get('Doc Index', 0) or 0)
                    pdf_viewer_url = result.get('PDF Viewer URL', '')
                    realestate_pdf_path = result.get('Real Estate PDF', '')

                    # Django ORM se database mein data save karo
                    obj, created = RealEstateData.objects.update_or_create(
                        search_name=search_name[:255],
                        entity_index=entity_index,
                        doc_index=doc_index,
                        defaults={
                            'pdf_viewer': pdf_viewer_url,
                            'realestate_pdf': realestate_pdf_path,
                        }
                    )

                    if created:
                        saved_count += 1
                        logger.debug(f"Saved new real estate record: {search_name}")

                except Exception as e:
                    logger.error(f"Error saving real estate result to DB: {e}")
                    logger.debug(f"Problematic result: {result}")
                    traceback.print_exc()

            logger.info(f"Successfully saved {saved_count} real estate records to database.")

            # Ab, database se data nikal kar Excel file mein save karo
            excel_path = scraper.save_results_to_excel()
            if excel_path:
                logger.info(f"SUCCESS: Real Estate data successfully saved to Excel at: {excel_path}")
            else:
                logger.error("FAILED: Failed to save Excel file from scraper results.")

        else:
            logger.warning("No real estate results found in scraper, nothing to save.")
        possible_locations = [OUTPUT_DIR, SCRAPERS_DIR, Path(".")]
        possible_patterns = ["realestate_index*", "realestate*", "RealEstate*"]
        
        latest_file = None
        for location in possible_locations:
            for pattern in possible_patterns:
                if location.exists():
                    files = list(location.glob(f"{pattern}.xlsx")) + list(location.glob(f"{pattern}.xls"))
                    if files:
                        latest_candidate = max(files, key=os.path.getmtime, default=None)
                        if latest_candidate and (latest_file is None or os.path.getmtime(latest_candidate) > os.path.getmtime(latest_file)):
                            latest_file = latest_candidate
        
        if latest_file:
            logger.info(f"Found real estate Excel file: {latest_file}")
            df = pd.read_excel(latest_file)
            logger.info(f"Excel file columns: {list(df.columns)}")
            
            saved_count = 0
            for _, row in df.iterrows():
                try:
                    row_data = {k: (str(v) if pd.notna(v) else '') for k, v in row.items()}
                    
                    obj, created = RealEstateData.objects.update_or_create(
                        search_name=row_data.get('search_name', '')[:255],
                        entity_index=int(row_data.get('entity_index', 0) or 0),
                        doc_index=int(row_data.get('doc_index', 0) or 0),
                        defaults={
                            'pdf_viewer': row_data.get('pdf_viewer', ''),
                            'realestate_pdf': row_data.get('final_url', ''),
                        }
                    )
                    
                    if created:
                        saved_count += 1
                        
                except Exception as e:
                    logger.error(f"Error saving row: {e}")
                    traceback.print_exc()

            logger.info(f"Saved {saved_count} real estate records from file")
        else:
            logger.warning("No real estate Excel file found")

    except Exception as e:
        logger.error(f"Error running real estate scraper: {e}")
        traceback.print_exc()
        
def find_latest_excel_file(directory, filename_prefix):
    """Find the latest Excel file with the given prefix"""
    try:
        # Look for both .xlsx and .xls files
        excel_files = list(directory.glob(f"{filename_prefix}*.xlsx")) + list(directory.glob(f"{filename_prefix}*.xls"))
        
        if not excel_files:
            return None
        
        # Return the most recently modified file
        return max(excel_files, key=os.path.getmtime)
    except Exception as e:
        logger.error(f"Error finding Excel file: {e}")
        return None
    
# ------------------ EXCEL DOWNLOAD VIEWS -------------------

@csrf_exempt
def download_lien_excel(request):
    """Download single lien record as Excel"""
    try:
        if request.method == 'POST':
            data = json.loads(request.body)
            pdf_url = data.get('pdf_url', '')
            
            # Find the lien record based on PDF URL
            lien_record = LienData.objects.filter(pdf_document_url=pdf_url).first()
            
            if not lien_record:
                return JsonResponse({'error': 'Record not found'}, status=404)
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Lien Record"
            
            # Add headers with styling
            headers = [
                'Direct Party (Debtor)', 'Reverse Party (Claimant)', 'Address', 
                'Zipcode', 'Total Due', 'County', 'Instrument', 'Date Filed',
                'Book', 'Page', 'Description', 'PDF Document URL', 'PDF File'
            ]
            
            # Style headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Add data
            data_row = [
                lien_record.direct_party_debtor or '',
                lien_record.reverse_party_claimant or '',
                lien_record.address or '',
                lien_record.zipcode or '',
                lien_record.total_due or '',
                lien_record.county or '',
                lien_record.instrument or '',
                lien_record.date_filed or '',
                lien_record.book or '',
                lien_record.page or '',
                lien_record.description or '',
                lien_record.pdf_document_url or '',
                lien_record.pdf_file or ''
            ]
            
            for col, value in enumerate(data_row, 1):
                ws.cell(row=2, column=col, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min((max_length + 2) * 1.2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"lien_record_{lien_record.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            wb.save(response)
            return response
            
        else:
            return JsonResponse({'error': 'Only POST method allowed'}, status=405)
            
    except Exception as e:
        logger.error(f"Error downloading lien Excel: {e}")
        return JsonResponse({'error': str(e)}, status=500)

def download_all_lien_excel(request):
    """Download all lien records as Excel"""
    try:
        # Get all lien records
        lien_records = LienData.objects.all().order_by('-created_at')
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "All Lien Data"
        
        # Add headers with styling
        headers = [
            'ID', 'Direct Party (Debtor)', 'Reverse Party (Claimant)', 'Address', 
            'Zipcode', 'Total Due', 'County', 'Instrument', 'Date Filed',
            'Book', 'Page', 'Description', 'PDF Document URL', 'PDF File', 'Created At'
        ]
        
        # Style headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data rows
        for row_num, record in enumerate(lien_records, 2):
            data_row = [
                record.id,
                record.direct_party_debtor or '',
                record.reverse_party_claimant or '',
                record.address or '',
                record.zipcode or '',
                record.total_due or '',
                record.county or '',
                record.instrument or '',
                record.date_filed or '',
                record.book or '',
                record.page or '',
                record.description or '',
                record.pdf_document_url or '',
                record.pdf_file or '',
                record.created_at.strftime('%Y-%m-%d %H:%M:%S') if record.created_at else ''
            ]
            
            for col, value in enumerate(data_row, 1):
                ws.cell(row=row_num, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2) * 1.2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"all_lien_data_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        logger.error(f"Error downloading all lien Excel: {e}")
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def download_realestate_excel(request):
    """Download real estate records as Excel based on search name"""
    try:
        if request.method == 'POST':
            data = json.loads(request.body)
            search_name = data.get('search_name', '')
            
            # Find real estate records based on search name
            if search_name:
                realestate_records = RealEstateData.objects.filter(search_name__icontains=search_name)
            else:
                realestate_records = RealEstateData.objects.all()
            
            if not realestate_records:
                return JsonResponse({'error': 'No records found'}, status=404)
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Real Estate Data"
            
            # Add headers with styling
            headers = ['Search Name', 'Entity Index', 'Document Index', 'PDF Viewer URL', 'Real Estate PDF URL', 'Created At']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2572a1", end_color="2572a1", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Add data rows
            for row_num, record in enumerate(realestate_records, 2):
                data_row = [
                    record.search_name or 'Not available',
                    record.entity_index or 0,
                    record.doc_index or 0,
                    record.pdf_viewer or 'Not available',
                    record.realestate_pdf or 'Not available',
                    record.created_at.strftime('%Y-%m-%d %H:%M:%S') if record.created_at else 'Not available'
                ]
                
                for col, value in enumerate(data_row, 1):
                    ws.cell(row=row_num, column=col, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min((max_length + 2) * 1.2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"realestate_{search_name or 'all'}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            wb.save(response)
            return response
            
        else:
            return JsonResponse({'error': 'Only POST method allowed'}, status=405)
            
    except Exception as e:
        logger.error(f"Error downloading real estate Excel: {e}")
        return JsonResponse({'error': str(e)}, status=500)

def download_all_realestate_excel(request):
    """Download all real estate records as Excel"""
    try:
        # Get all real estate records
        realestate_records = RealEstateData.objects.all().order_by('-created_at')
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "All Real Estate Data"
        
        # Add headers with styling
        headers = [
            'ID', 'Search Name', 'Entity Index', 'Document Index', 
            'PDF Viewer URL', 'Real Estate PDF URL', 'Created At'
        ]
        
        # Style headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2572a1", end_color="2572a1", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data rows
        for row_num, record in enumerate(realestate_records, 2):
            data_row = [
                record.id,
                record.search_name or '',
                record.entity_index or '',
                record.doc_index or '',
                record.pdf_viewer or '',
                record.realestate_pdf or '',
                record.created_at.strftime('%Y-%m-%d %H:%M:%S') if record.created_at else ''
            ]
            
            for col, value in enumerate(data_row, 1):
                ws.cell(row=row_num, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2) * 1.2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"all_realestate_data_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        logger.error(f"Error downloading all real estate Excel: {e}")
        return JsonResponse({'error': str(e)}, status=500)