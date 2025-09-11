import os
import re
import json
import math
import random
import asyncio
import traceback
import pandas as pd
import requests
import openpyxl
import img2pdf
import aiohttp
import ssl
import certifi
import pytesseract
import cv2
from fuzzywuzzy import fuzz
import numpy as np
from datetime import datetime
from pathlib import Path
from PIL import Image
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from rich.console import Console
import playwright.async_api as pw
from typing import Tuple, Optional, Any, Dict
from playwright.async_api import TimeoutError as PlaywrightTimeoutError
from fuzzywuzzy import fuzz

# --- Constants for better readability ---
MAX_RP_TO_PROCESS = 5
TYPING_DELAY_MIN = 100
TYPING_DELAY_MAX = 250
OCR_KEYWORD = "TOTAL DUE"
FUZZY_MATCH_THRESHOLD = 80 # A higher number means a stricter match

load_dotenv()
console = Console()

if os.name == "nt": 
    pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

# ---------- config -------------------------------------------------------------
HEADLESS = False 
STATE_FILE = Path("cookies.json")
TAX_EMAIL = os.getenv("GSCCCA_USERNAME")
TAX_PASSWORD = os.getenv("GSCCCA_PASSWORD")
LOCALE = "en-GB"
TIMEZONE = "UTC"
VIEWPORT = {"width": 1366, "height": 900}
UA_DICT = {
    "mac": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36",
    "linux": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36",
    "win": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome"
}
UA_TYPE = "win"
UA = UA_DICT.get(UA_TYPE, UA_DICT["win"])
EXTRA_HEADERS = {
    "Accept-Language": "en-GB,en-US;q=0.9,en;q=0.8"
}

# ---------- core scraping ----------------------------------------------------
class GSCCCAScraper:
    """Scrape the latest tax records from GSCCCA pages."""

    def __init__(self) -> None:
        self.page = None
        self.email = TAX_EMAIL
        self.password = TAX_PASSWORD
        self.homepage = "https://www.gsccca.org/"
        self.login_url = "https://apps.gsccca.org/login.asp"
        self.name_search_url = "https://search.gsccca.org/Lien/namesearch.asp"
        self.results = []
        self.ssl_context = ssl.create_default_context(cafile=certifi.where())
        

    def time_sleep(self, a: int = 2500, b: int = 5000) -> int:
        return random.uniform(a, b)

    async def dump_cookies(self, out_file="cookies.json"):
        """Save cookies + storage ONLY for login check."""
        try:
            state = await self.page.context.storage_state()
            Path(out_file).write_text(json.dumps(state, indent=2))
            print(f"Saved login state to --> {out_file}")
        except Exception as e:
            console.print(f"[red]Failed to dump cookies: {e}[/red]")

    async def already_logged_in(self) -> bool:
        """Check if user is already logged in."""
        try:
            await self.page.wait_for_load_state("domcontentloaded")
            all_text = await self.page.evaluate("document.body.innerText")
            if "logout" in all_text.lower():
                print("Logout link detected -> User is already logged in.")
                return True

            print("Could not detect login/logout clearly.")
            return False

        except Exception as e:
            print(f"[already_logged_in ERROR] {e}")
            return False

    async def login_(self, email: str, password: str):
        """Perform login and save cookies."""
        print("[LOGIN] Navigating to login page...")
        await self.page.goto(self.login_url, wait_until="domcontentloaded", timeout=60000)
        await self.page.wait_for_timeout(self.time_sleep())
        await self.check_and_handle_announcement()

        print(f"[LOGIN] Filling username: {email}")
        await self.page.fill("input[name='txtUserID']", email)
        await self.page.wait_for_timeout(2000)

        print("[LOGIN] Filling password...")
        await self.page.fill("input[name='txtPassword']", password)
        await self.page.wait_for_timeout(2000) 

        print("[LOGIN] Checking 'Remember login details' checkbox if not already checked...")
        checkbox = await self.page.query_selector("input[type='checkbox'][name='permanent']")
        if checkbox:
            is_checked = await checkbox.is_checked()
            if not is_checked:
                await checkbox.click()
                print("[LOGIN] Checkbox clicked.")
            else:
                print("[LOGIN] Checkbox already checked.")
        else:
            print("[LOGIN] Checkbox not found on the page.")

        try:
            print("[LOGIN] Clicking login button...")
            await self.page.click("img[name='logon']")
        except Exception as e:
            print(f"[LOGIN] Click failed: {e}, using JS submit...")
            await self.page.evaluate("document.forms['frmLogin'].submit()")

        await self.page.wait_for_load_state("networkidle", timeout=15000)

        if await self.page.query_selector("a:has-text('Logout')"):
            print("[LOGIN] Login successful!")
            await self.dump_cookies()
            return True
        else:
            print("[LOGIN] Login failed")
            return False

    async def step1_open_homepage(self):
        """Go to homepage & check if logged in."""
        print("[STEP 1] Opening homepage...")
        await self.page.goto(self.homepage, wait_until="domcontentloaded")
        await self.page.wait_for_timeout(self.time_sleep())

        if await self.already_logged_in():
            print("[STEP 1] Logged in detected")
            return True
        else:
            print("[STEP 1] Not logged in")
            return False
        
    async def check_and_handle_announcement(self):
        """Check if announcement page loaded; if yes, redirect to name_search_url."""
        current_url = self.page.url
        if "CustomerCommunicationApiAnnouncement1.asp" in current_url:
            print("[INFO] Announcement page detected. Redirecting to name search...")
            await self.page.select_option("#Options", "dismiss")
            await self.page.wait_for_timeout(1500)
            await self.page.click("input[name='Continue']")

    async def step2_click_name_search(self):
        """Go directly to Name Search page."""
        print("[STEP 2] Going directly to Name Search page...")
        try:
            await self.page.goto(self.name_search_url, wait_until="domcontentloaded", timeout=60000)
            await self.page.wait_for_timeout(self.time_sleep())
            await self.check_and_handle_announcement()
            print("[STEP 2] Landed on Name Search page")
        except Exception as e:
            print(f"[STEP 2 ERROR] {e}")

    async def step3_fill_form(self):
        """Fill Name Search form with given details."""
        print("[STEP 3] Filling Name Search form...")

        try:
            await self.page.select_option("#txtPartyType", "2")
            await self.page.wait_for_timeout(self.time_sleep())

            await self.page.select_option("select[name='txtInstrCode']", "2")
            await self.page.wait_for_timeout(self.time_sleep())

            await self.page.select_option("select[name='intCountyID']", "64")
            await self.page.wait_for_timeout(self.time_sleep())

            await self.page.check("input[name='bolInclude'][value='0']")
            await self.page.wait_for_timeout(self.time_sleep())

            search_box = await self.page.query_selector("#txtSearchName")
            await search_box.click()
            for ch in "gordon":
                await self.page.keyboard.type(ch, delay=random.randint(TYPING_DELAY_MIN, TYPING_DELAY_MAX)) 
            await self.page.wait_for_timeout(self.time_sleep())

            await self.page.fill("input[name='txtFromDate']", "")
            for ch in "01/01/2025":
                await self.page.keyboard.type(ch, delay=random.randint(TYPING_DELAY_MIN, TYPING_DELAY_MAX))

            await self.page.fill("input[name='txtToDate']", "")
            for ch in "09/23/2025":
                await self.page.keyboard.type(ch, delay=random.randint(TYPING_DELAY_MIN, TYPING_DELAY_MAX))
            await self.page.wait_for_timeout(self.time_sleep())

            await self.page.select_option("select[name='MaxRows']", "100")
            await self.page.wait_for_timeout(self.time_sleep())

            await self.page.select_option("select[name='TableType']", "1")
            await self.page.wait_for_timeout(self.time_sleep())
        
            await self.page.click("form[name='SearchType'] input[value='Search']")
            print("[STEP 3] Form filled successfully")

        except Exception as e:
            print(f"[STEP 3 ERROR] {e}")

    async def step4_select_highest_occurs(self):
        """On liennames.asp page, select row with highest Occurs score."""
        print("[STEP 4] Selecting row with highest Occurs...")

        try:
            await self.page.wait_for_selector("table.name_results", timeout=30_000)
            rows = await self.page.query_selector_all("table.name_results tr")
            highest_occurs = -1
            best_radio = None
            for row in rows[1:]:  
                cols = await row.query_selector_all("td")
                if len(cols) < 3:
                    continue
                occurs_text = await cols[1].inner_text()
                name_text = await cols[2].inner_text()
                radio = await cols[0].query_selector("input[type='radio']")
                try:
                    occurs = int(occurs_text.strip())
                except:
                    continue

                if occurs > highest_occurs:
                    highest_occurs = occurs
                    best_radio = radio

            if best_radio:
                await best_radio.click()
                print(f"[STEP 4] Selected row with highest Occurs = {highest_occurs}")
            else:
                print("[STEP 4] No rows found to select")
                
            await self.page.click("input[value='Display Details']")
            await self.page.wait_for_load_state("domcontentloaded")
            await self.page.wait_for_timeout(self.time_sleep())

        except Exception as e:
            print(f"[STEP 4 ERROR] {e}")

    async def human_delay(self, min_t=0.8, max_t=2.0):
        t = random.uniform(min_t, max_t)
        await asyncio.sleep(t)

    async def human_scroll(self, min_y=200, max_y=800):
        y = random.randint(min_y, max_y)
        await self.page.mouse.wheel(0, y)
        print(f"[HUMAN] Scrolled {y}px")
        await self.human_delay(0.5, 1.2)

    async def process_rp_details(self):
        """ Step 5: Process all RP buttons, extract data and save """
        self.results = []
        visited_pages = set()

        while True:
            rp_links = await self.page.query_selector_all("a[href*='lienfinal']")
            if not rp_links:
                print("[WARNING] No RP buttons found on this page")
                break

            first_link = await rp_links[0].get_attribute("href")
            if first_link in visited_pages:
                print(f"[INFO] Duplicate page detected → already visited. Stopping loop.")
                break
            visited_pages.add(first_link)

            total = len(rp_links)
            print(f"[INFO] Found {total} RP buttons on this page")

            for i in range(min(MAX_RP_TO_PROCESS, total)): 
                try:
                    current_links = await self.page.query_selector_all("a[href*='lienfinal']")
                    if i >= len(current_links):
                        continue

                    link = current_links[i]

                    retries = 3
                    for attempt in range(retries):
                        try:
                            await link.click()
                            await self.page.wait_for_load_state("domcontentloaded")
                            break
                        except Exception as e:
                            print(f"[ERROR] Click failed (Attempt {attempt+1}/{retries}): {e}")
                            if attempt == retries - 1:
                                raise
                            await asyncio.sleep(3)

                    data = await self.parse_rp_detail()
                    if data:
                        self.results.append(data)
                        print(f"[SUCCESS] Saved RP index {i+1} → "
                                f"{data.get('Name Selected','N/A')} | "
                                f"Book={data.get('Book','')} Page={data.get('Page','')}")
                        pd.DataFrame(self.results).to_excel("LienResults.xlsx", index=False)

                    await asyncio.sleep(2)
                    back_btn = await self.page.query_selector("input[name='bBack']")
                    if back_btn:
                        await back_btn.click()
                        await self.page.wait_for_load_state("domcontentloaded")
                    else:
                        await self.page.go_back()
                        await self.page.wait_for_load_state("domcontentloaded")

                except Exception as e:
                    print(f"[ERROR] Failed at RP index {i}: {e}")
                    try:
                        await self.page.go_back()
                        await self.page.wait_for_load_state("domcontentloaded")
                    except:
                        pass
                    continue

            next_page = await self.page.query_selector("a:has-text('Next')")
            if next_page:
                print("[INFO] Going to next page...")
                try:
                    await next_page.click()
                    await self.page.wait_for_load_state("domcontentloaded")
                except Exception as e:
                    print(f"[ERROR] Pagination failed: {e}")
                    break
            else:
                print("[INFO] No more pages found.")
                break

        self.save_to_excel("LienResults.xlsx")
        
    async def parse_rp_detail(self):
        """ Helper: Parse lienfinal.asp detail page with BeautifulSoup + Viewer URL + Single Page PDF + OCR + Address2 + Zipcode2 + Total Due """
        await self.page.wait_for_load_state("domcontentloaded", timeout=15000)
        await asyncio.sleep(1.5)
        html = await self.page.content()
        soup = BeautifulSoup(html, "html.parser")
        data = {}

        def safe_text(el):
            return el.get_text(" ", strip=True) if el else ""

        def extract_total_due(text: str) -> str:
            text = text.replace('\n', ' ').replace('\r', ' ')

            def find_fuzzy_match_and_extract_amount(search_text, keyword):
                lines = [line.strip() for line in search_text.splitlines() if line.strip()]
                for i, line in enumerate(lines):
                    if fuzz.ratio(line.upper(), keyword.upper()) >= FUZZY_MATCH_THRESHOLD:
                        print(f"Fuzzy match found: '{line}' for '{keyword}'")
                        amounts = re.findall(r"[\d,]+\.\d{2}", line)
                        if amounts:
                            return amounts[-1]
                        
                        # Look in the next line if no amount found in the current line
                        if i + 1 < len(lines):
                            next_line = lines[i+1]
                            amounts_in_next_line = re.findall(r"[\d,]+\.\d{2}", next_line)
                            if amounts_in_next_line:
                                return amounts_in_next_line[-1]
                return None
            
            # Strategy 1: Direct keyword search
            m1 = re.search(r"TOTAL\s*DUE(?:\s*[:\s-]*\s*)([\d,]+\.\d{2})", text, re.IGNORECASE)
            m2 = re.search(r"TOT(?:AL)?\s*DUE\s*\$?\s*([\d,]+\.\d{2})", text, re.IGNORECASE)
            
            if m1:
                return f"{float(m1.group(1).replace(',', '')):.2f}"
            if m2:
                return f"{float(m2.group(1).replace(',', '')):.2f}"

            # Strategy 2: Calculate from individual tax items
            print("TOTAL DUE keyword not found. Attempting to calculate from tax items.")
            amounts = {}
            patterns = {
                "TAX": r"TAX\s*\$?\s*([\d,]+\.\d{2})",
                "PENALTY": r"PENALTY\s*\$?\s*([\d,]+\.\d{2})",
                "FIFA": r"FIFA\s*\$?\s*([\d,]+\.\d{2})",
                "GED": r"GED\s*\$?\s*([\d,]+\.\d{2})",
                "INTEREST": r"INTEREST\s*\$?\s*([\d,]+\.\d{2})",
                "DEMO LIEN": r"DEMO\s+LIEN\s*\$?\s*([\d,]+\.\d{2})",
                "PAYMENT": r"PAYMENT\(S\)\s*\$?\s*([\d,]+\.\d{2})"
            }

            total = 0.0
            found_items = False
            for key, pattern in patterns.items():
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    found_items = True
                    try:
                        amount = float(match.group(1).replace(",", ""))
                        amounts[key] = amount
                        print(f"Found {key}: {amount:.2f}")
                        if key != "PAYMENT":
                            total += amount
                        else:
                            total -= amount
                    except ValueError:
                        print(f"Failed to parse amount for {key}")

            if found_items:
                print(f"Calculated Total Due: {total:.2f}")
                return f"{total:.2f}"
            
            # Strategy 3: Fuzzy keyword matching
            print("Calculation from tax items failed. Trying fuzzy matching.")
            fuzzy_amount_str = find_fuzzy_match_and_extract_amount(text, OCR_KEYWORD)
            if fuzzy_amount_str:
                try:
                    return f"{float(fuzzy_amount_str.replace(',', '')):.2f}"
                except ValueError:
                    pass

            print("WARNING: Could not find TOTAL DUE or calculate from tax items.")
            return "Not Found"

        def extract_addresses_from_ocr(text, max_addresses=2):
            lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
            addresses = []
            for idx, ln in enumerate(lines):
                m = re.search(r'([A-Za-z][A-Za-z0-9\.\'&\-\s]+,\s*[A-Za-z]{2}\s*\d{5})', ln)
                if m:
                    city_state_zip = m.group(1).strip()
                    street = ""
                    for j in range(1, 6):
                        if idx - j < 0:
                            break
                        prev = lines[idx - j]
                        if re.search(r'\b(County|Tax|Commissioner|Recorded|Doc:|Rept#|VS\b|Defendant|GRANT|PAYMENT|TOTAL DUE|PHONE|TEL|Fax)\b', prev, re.I):
                            continue
                        if re.search(r'^\d+\s', prev) or re.search(r'\b(St(reet)?|Street|Rd|Road|Ave|Avenue|Blvd|Ln|Lane|Dr|Drive|Way|Ct|Court|PKWY|Parkway)\b', prev, re.I):
                            street = prev
                            break
                        if not re.search(r'^\b(Grant|GORDON|LIEN|TOTAL|PAYMENT)\b', prev, re.I):
                            street = prev
                            break
                    full_address = (street + " " + city_state_zip).strip() if street else city_state_zip
                    zip_m = re.search(r'(\d{5})$', city_state_zip)
                    zipcode = zip_m.group(1) if zip_m else ""
                    addresses.append({"address": full_address, "zipcode": zipcode})
                    if len(addresses) >= max_addresses:
                        break
            while len(addresses) < max_addresses:
                addresses.append({"address": "", "zipcode": ""})
            return addresses
        
        def deskew_image(image):
            coords = np.column_stack(np.where(image > 0))
            if coords.size == 0: 
                return image
            
            angle = cv2.minAreaRect(coords)[-1]
            if angle < -45:
                angle = -(90 + angle)
            else:
                angle = -angle
            
            (h, w) = image.shape[:2]
            center = (w // 2, h // 2)
            M = cv2.getRotationMatrix2D(center, angle, 1.0)
            rotated = cv2.warpAffine(image, M, (w, h), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)
            return rotated

        def get_robust_ocr_text(image_path: str) -> str:
            """
            Applies a prioritized, intelligent fallback strategy with various image
            processing techniques to get the best OCR text, focusing on efficiency.
            """
            img = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
            
            deskewed_img = deskew_image(img)

            processing_methods = {
                "Original": lambda x: x,
                "Inverted": lambda x: cv2.bitwise_not(x),
                "Adaptive Threshold": lambda x: cv2.adaptiveThreshold(x, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2),
                "Otsu Threshold": lambda x: cv2.threshold(x, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1],
                "CLAHE": lambda x: cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8)).apply(x),
                "Denoised": lambda x: cv2.fastNlMeansDenoising(x, None, 30, 7, 21),
                "Sharpened": lambda x: cv2.filter2D(x, -1, kernel=np.array([[-1, -1, -1], [-1, 9, -1], [-1, -1, -1]])),
                "Contrast Enhanced": lambda x: cv2.convertScaleAbs(x, alpha=1.5, beta=0),
                "Median Blurred": lambda x: cv2.medianBlur(x, 3),
                "Unsharp Masked": lambda x: cv2.addWeighted(x, 1.5, cv2.GaussianBlur(x, (0, 0), 3), -0.5, 0),
            }

            prioritized_methods = [
                "Original",
                "Inverted",
                "Adaptive Threshold",
                "Otsu Threshold"
            ]

            print("Trying prioritized OCR methods on deskewed image...")
            for method_name in prioritized_methods:
                try:
                    processed_img = processing_methods[method_name](deskewed_img)
                    text = pytesseract.image_to_string(Image.fromarray(processed_img), lang="eng")
                    print(f"OCR Text ({method_name} on Deskewed):\n{text.strip()}")
                    if "TOTAL DUE" in text.upper():
                        print(f"SUCCESS: '{OCR_KEYWORD}' found with '{method_name}' method on deskewed image.")
                        return text.strip()
                except Exception as e:
                    print(f"Error with {method_name}: {e}")
                    continue

            print(f"'{OCR_KEYWORD}' not found. Falling back to other methods...")
            for method_name, method_func in processing_methods.items():
                if method_name in prioritized_methods:
                    continue
                try:
                    processed_img = method_func(deskewed_img)
                    text = pytesseract.image_to_string(Image.fromarray(processed_img), lang="eng")
                    print(f"OCR Text (Fallback: {method_name} on Deskewed):\n{text.strip()}")
                    if "TOTAL DUE" in text.upper():
                        print(f"SUCCESS: '{OCR_KEYWORD}' found with fallback '{method_name}' method.")
                        return text.strip()
                except Exception as e:
                    print(f"Error with fallback {method_name}: {e}")
                    continue

            print(f"WARNING: '{OCR_KEYWORD}' not found in any method. Returning empty string.")
            return ""

        # ---------- Normal Data Extraction from HTML ----------
        header_table = soup.find("table", cellpadding="2")
        if header_table:
            rows = header_table.find_all("tr")
            for row in rows:
                cells = row.find_all("td")
                if len(cells) == 2:
                    key = safe_text(cells[0]).rstrip(":")
                    val = safe_text(cells[1])
                    data[key] = val
                elif len(cells) == 1:
                    data["Extra Info"] = safe_text(cells[0])

        doc_table = soup.find("table", width="800", cellpadding="0", cellspacing="0")
        if doc_table:
            rows = doc_table.find_all("tr")[1:]
            if rows:
                cols = [safe_text(td) for td in rows[0].find_all("td")]
                if len(cols) >= 6:
                    data.update({
                        "County": cols[0],
                        "Instrument": cols[1],
                        "Date Filed": cols[2],
                        "Time": cols[3],
                        "Book": cols[4],
                        "Page": cols[5],
                    })

        desc_table = soup.find("td", string=lambda t: t and "Description" in t)
        if desc_table:
            tbody = desc_table.find_parent("table")
            desc_val = safe_text(tbody.find_all("tr")[1].find("td"))
            data["Description"] = desc_val

        debtor_table = soup.find("td", string=lambda t: t and "Direct Party (Debtor)" in t)
        if debtor_table:
            tbody = debtor_table.find_parent("table")
            debtors = [safe_text(td) for td in tbody.find_all("td")[1:]]
            data["Direct Party (Debtor)"] = "; ".join(debtors)

        claimant_table = soup.find("td", string=lambda t: t and "Reverse Party (Claimant)" in t)
        if claimant_table:
            tbody = claimant_table.find_parent("table")
            claimants = [safe_text(td) for td in tbody.find_all("td")[1:]]
            data["Reverse Party (Claimant)"] = "; ".join(claimants)

        cross_table = soup.find("td", string=lambda t: t and "Cross-Referenced Instruments" in t)
        if cross_table:
            tbody = cross_table.find_parent("table")
            rows = tbody.find_all("tr")[1:]
            refs = []
            for row in rows:
                cols = [safe_text(td) for td in row.find_all("td")]
                if any(cols):
                    refs.append(" | ".join(cols))
            data["Cross-Referenced Instruments"] = "; ".join(refs)

        record_info = soup.find("i")
        if record_info:
            data["Record Added"] = safe_text(record_info)

        # ---------- Updated PDF & OCR Logic ----------
        viewer_script = soup.find("script", string=lambda t: t and "ViewImage" in t)
        if viewer_script:
            try:
                script_text = viewer_script.string
                match = re.search(r'var iLienID\s*=\s*(\d+);', script_text)
                if not match:
                    raise ValueError("Lien ID not found in script.")
                
                lien_id = match.group(1)
                county = re.search(r'var county\s*=\s*"(\d+)"', script_text).group(1)
                book = re.search(r'var book\s*=\s*"(\d+)"', script_text).group(1)
                page_num = re.search(r'var page\s*=\s*"(\d+)"', script_text).group(1)
                userid = re.search(r'var user\s*=\s*(\d+)', script_text).group(1)
                appid = re.search(r'var appid\s*=\s*(\d+)', script_text).group(1)

                viewer_url = (f"https://search.gsccca.org/Imaging/HTML5Viewer.aspx?id={lien_id}&key1={book}&key2={page_num}&county={county}&userid={userid}&appid={appid}")
                data["PDF Document URL"] = viewer_url

                debtor_name = data.get("Direct Party (Debtor)", "UnknownDebtor").split(";")[0][:40].replace(" ", "_").replace(",", "")
                pdf_name = f"{debtor_name}_Page{page_num}.pdf"
                download_dir = "downloads"
                os.makedirs(download_dir, exist_ok=True)
                pdf_path = os.path.join(download_dir, pdf_name)
                tmp_img = os.path.join(download_dir, f"tmp_{page_num}.png")

                popup = await self.page.context.new_page()
                await popup.goto(viewer_url, timeout=30000)
                await popup.wait_for_load_state("domcontentloaded")
                await asyncio.sleep(2)
                await popup.wait_for_selector("div.vtm_imageClipper canvas", timeout=10000)
                canvas = await popup.query_selector("div.vtm_imageClipper canvas")

                if canvas:
                    await canvas.screenshot(path=tmp_img)
                    data["PDF"] = pdf_name
                
                await popup.close()
                
                try:
                    text = get_robust_ocr_text(tmp_img)
                    data["OCR_Text"] = text
                    
                    if text:
                        addr_list = extract_addresses_from_ocr(data["OCR_Text"], max_addresses=2)
                        data["Address2"] = addr_list[1]["address"]
                        data["Zipcode2"] = addr_list[1]["zipcode"]
                        data["Total Due"] = extract_total_due(data["OCR_Text"])
                    else:
                        data["OCR_Text"] = ""
                        data["Address2"] = ""
                        data["Zipcode2"] = ""
                        data["Total Due"] = ""
                except Exception as e:
                    print(f"[ERROR] Failed to process OCR and extract data: {e}")
                    data["OCR_Text"] = ""
                    data["Address2"] = ""
                    data["Zipcode2"] = ""
                    data["Total Due"] = ""
                
                if os.path.exists(tmp_img):
                    os.remove(tmp_img)

            except Exception as e:
                print(f"[ERROR] An issue occurred during PDF/OCR processing: {e}")
                data.update({
                    "PDF Document URL": "",
                    "PDF": "",
                    "OCR_Text": "",
                    "Address2": "",
                    "Zipcode2": "",
                    "Total Due": ""
                })

        return data
        
    def save_to_excel(self, filename="LienResults.xlsx"):
        if not hasattr(self, "results") or not self.results:
            print("[WARNING] No results to save")
            return

        df = pd.DataFrame(self.results)
        columns = [
            "Direct Party (Debtor)", "Reverse Party (Claimant)",
            "Address2", "Zipcode2", "Total Due",
            "Name Selected", "Searched", "User Selected Dates", "County Good From", "Query Made",
            "County", "Instrument", "Date Filed", "Time", "Book", "Page",
            "Description", "Sec/GMD", "District", "Land Lot", "Subdivision",
            "Unit", "Block", "Lot", "Comment",
            "Cross-Referenced Instruments", "Record Added",
            "PDF Document URL", "PDF", "OCR_Text",
        ]
        for col in columns:
            if col not in df.columns:
                df[col] = ""
        df = df[columns]

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        base, ext = os.path.splitext(filename)
        final_filename = f"{base}_{ts}{ext}"
        df.to_excel(final_filename, index=False)
        
        import openpyxl
        wb = openpyxl.load_workbook(final_filename)
        ws = wb.active
        headers = {cell.value: cell.column for cell in ws[1]}
        pdf_col = headers.get("PDF")
        if pdf_col:
            download_dir = "downloads"
            for row in range(2, ws.max_row + 1):
                pdf_name = ws.cell(row=row, column=pdf_col).value
                if pdf_name:
                    pdf_path = os.path.abspath(os.path.join(download_dir, pdf_name))
                    ws.cell(row=row, column=pdf_col).value = f'=HYPERLINK("{pdf_path}", "{pdf_name}")'
        wb.save(final_filename)
        print(f"[INFO] Saved {len(df)} records to {final_filename} (with PDF links)")

    async def scrape(self, page_url: str) -> Dict[str, Any]:
        """Open *url* in the current page and return latest post data."""
        try:
            await self.page.goto(page_url, wait_until="domcontentloaded", timeout=60_000)
            await self.page.wait_for_timeout(3000)
            title = await self.page.title()
            rows = await self.page.query_selector_all("table tr")
            data = []
            for row in rows:
                text = await row.inner_text()
                data.append(text.strip())
            return {
                "page_title": title,
                "row_count": len(rows),
                "rows": data[:10]
            }
        except Exception as e:
            console.print(f"[red]Scrape error: {e}[/red]")
            return {}

    async def run(self):
        """initialize playwright browser and run scraper"""
        playwright = await pw.async_playwright().start()
        browser = await playwright.chromium.launch(
            headless=HEADLESS, 
            channel="chrome",
            args=[
                "--disable-blink-features=AutomationControlled",
                "--start-maximized",
            ]
        )
        if STATE_FILE.exists():
            print("Loaded session from storage...")
            context = await browser.new_context(
                storage_state=STATE_FILE,
                user_agent=UA,
                locale=LOCALE,
                timezone_id=TIMEZONE,
                viewport=VIEWPORT,
                device_scale_factor=1,
                extra_http_headers=EXTRA_HEADERS,
            )
            self.page = await context.new_page()
        else:
            print("Starting fresh session...")
            context = await browser.new_context(
                user_agent=UA,
                locale=LOCALE,
                timezone_id=TIMEZONE,
                viewport=VIEWPORT,
                device_scale_factor=1,
                extra_http_headers=EXTRA_HEADERS,
            )
            self.page = await browser.new_page()
        if STATE_FILE.exists():
            print("Loaded session from cookies.json...")
            await context.add_cookies(json.loads(Path(STATE_FILE).read_text())["cookies"])
            await self.page.goto(self.homepage, wait_until="domcontentloaded")
        else:
            print("Starting fresh login...")
            await self.page.goto(self.login_url, wait_until="domcontentloaded")
            if not await self.already_logged_in():
                if await self.login_(email=self.email, password=self.password):
                    await self.page.wait_for_timeout(self.time_sleep())
                    await self.dump_cookies()
        if not await self.step1_open_homepage():
            await self.login_(email=self.email, password=self.password)
            await self.page.wait_for_timeout(self.time_sleep())
        await self.step2_click_name_search()
        await self.check_and_handle_announcement()
        await self.step3_fill_form()
        await self.check_and_handle_announcement()
        await self.step4_select_highest_occurs()
        await self.check_and_handle_announcement()
        await self.process_rp_details()  
        await self.check_and_handle_announcement()
        self.save_to_excel()
        await browser.close()
        await playwright.stop()

async def main() -> None:
    scraper = GSCCCAScraper()
    await scraper.run()

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except Exception as e:
        console.print("\n[bold yellow]Interrupted![/bold yellow]\n", traceback.format_exc())
    finally:
        console.print("[bold green]Exiting...[/bold green]")